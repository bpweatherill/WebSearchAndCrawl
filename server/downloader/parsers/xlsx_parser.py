from openpyxl import load_workbook
from typing import Optional

def parse_xlsx(filepath: str) -> Optional[str]:
    """Extract text from an XLSX file."""
    try:
        wb = load_workbook(filepath, read_only=True)
        text = []
        for sheet in wb:
            for row in sheet.iter_rows(values_only=True):
                text.append("\t".join(str(cell) for cell in row if cell is not None))
        return "\n".join(text)
    except Exception:
        return None
